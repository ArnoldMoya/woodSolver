# -*- coding: utf-8 -*-
"""
Created on Tue Aug  2 21:16:25 2022

@author: Lenovo
"""
from ortools.linear_solver import pywraplp
import openpyxl 
from collections import defaultdict

def newSolver(name,integer=False):
  return pywraplp.Solver(name,\
                         pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING \
                         if integer else \
                         pywraplp.Solver.GLOP_LINEAR_PROGRAMMING)

'''
return a printable value
'''
def SolVal(x):
  if type(x) is not list:
    return 0 if x is None \
      else x if isinstance(x,(int,float)) \
           else x.SolutionValue() if x.Integer() is False \
                else int(x.SolutionValue())
  elif type(x) is list:
    return [SolVal(e) for e in x]

def ObjVal(x):
  return x.Objective().Value()

def countRows(sheet, col):
    i = 1
    while (sheet.cell(row = i, column = col).value):
        #cell_obj = sheet.cell(row = i, column = col)
        #print(cell_obj.value)
        i = i + 1
    return i

def readFile(path):
    file = openpyxl .load_workbook(path) 
    sheet = file.active 

    # count columns   
    n_profile = countRows(sheet, 1)
    n_kind = countRows(sheet, 2)
    n_quantity = countRows(sheet, 3)
    n_lenght = countRows(sheet, 4)
    
    n = [n_profile, n_kind, n_quantity, n_lenght]
    
    # status
        # 0: optimal
        # 1: feasible
        # 2: infeasible
        # 3: unbounded
        # 4: abnormal
        # 6: notsolved
        # 7: not stock
        # 8: error in file
    if max(n) > min(n): 
        status = 8  # 
        return status, {} 

    def def_value():
        return []
    description = defaultdict(def_value)
    
    for i in range(2, n_profile):
        wood = sheet.cell(row = i, column = 1).value
        kind = sheet.cell(row = i, column = 2).value
        quantity = sheet.cell(row = i, column = 3).value
        lenght = sheet.cell(row = i, column = 4).value
        #print(wood, kind, quantity, lenght)
        for j in range(quantity):
            description[(wood, kind)].append(int(lenght))
    
    status = 0 #Everything fine
    return status, description