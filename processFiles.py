# -*- coding: utf-8 -*-
"""
Created on Wed Aug  3 12:27:00 2022

@author: Lenovo
"""

import openpyxl 
from solverAllCases import *
from utils import *
from collections import defaultdict

def processFiles(requirement_path, stock_path, output_path):
    #requirement_path = 'D:/Fiver/alikbraz/woodSolver/requirements.xlsx'
    #stock_path = 'D:/Fiver/alikbraz/woodSolver/stock.xlsx'
    # Read fies
    requirement = readFile(requirement_path)
    stock = readFile(stock_path)
    
    if requirement[0] > 0:
        print("error reading requirement")
    
    if stock[0] > 0:
        print("error reading stock")
       
    # solve problem
    strategy = solverAllCases(requirement[1], stock[1])
    
    # write output
    wb = openpyxl.Workbook()
    sheet = wb.active
    #titles
    sheet.cell(row = 1, column = 1).value = "Wood profile"
    sheet.cell(row = 1, column = 2).value = "Wood kind"
    sheet.cell(row = 1, column = 3).value = "Status"
    sheet.cell(row = 1, column = 4).value = "Stock length"
    sheet.cell(row = 1, column = 5).value = "Cuts"
    
    i=2
    for material, solution in strategy.items():
        for j in range(len(solution[1])):
            c = sheet.cell(row = i, column = 1)
            c.value = str(material[0])
            
            c = sheet.cell(row = i, column = 2)
            c.value = str(material[1])
            
            c = sheet.cell(row = i, column = 3)
            if solution[0] == 0:
                c.value = "Solved"
            else:
                c.value = "Not Solved"
            
            c = sheet.cell(row = i, column = 4)
            c.value = solution[1][j][0]
            
            c = sheet.cell(row = i, column = 5)
            c.value = str(solution[1][j][1])
        
            i = i+1
    wb.save(output_path) #"D:/Fiver/alikbraz/woodSolver/solve.xlsx"